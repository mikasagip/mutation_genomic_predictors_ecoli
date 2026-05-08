from collections import defaultdict
import re
import glob
import time
import math
import numpy as np
import matplotlib.pyplot as plt
from scipy import stats
from scipy.stats import wilcoxon
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


strt_time = time.time()

rates_foster_pos1 = open("rates_foster_pos1.csv", "w")
rates_foster_pos2 = open("rates_foster_pos2.csv", "w")
rates_foster_pos3 = open("rates_foster_pos3.csv", "w")

rates_wei_pos1 = open("rates_wei_pos1.csv", "w")
rates_wei_pos2 = open("rates_wei_pos2.csv", "w")
rates_wei_pos3 = open("rates_wei_pos3.csv", "w")
#creates outfiles for vectors

wb = Workbook()
ws = wb.active
ws.title = "formatted_rates"
#activate excel file
headers = ["Trinucleotide", "Foster P1 (x10³)", "Foster P2 (x10³)", "Foster P3 (x10³)", "Wei P1 (x10³)", "Wei P2 (x10³)", "Wei P3 (x10³)"]
#write headers
ws.append(headers)
#append the headers to the file
for cell in ws[1]:
    cell.font = Font(bold=True)

rates_f = [rates_foster_pos1, rates_foster_pos2, rates_foster_pos3, rates_wei_pos1,rates_wei_pos2, rates_wei_pos3]

for f in rates_f:
    f.write("trinucleotide,count_mut,count_tri,mutation_rate\n")
    #writes titles in all of the vector files


out_f = open("selection_test_results_py.csv", "w")
#creates and opens the stats out file in edit mode 

out_f.write(f"dataset,diff_pos,v_statistic,p_value\n")

genes_f = glob.glob("genes*.csv")
seq_f = glob.glob("seq*.fna")
mut_f =glob.glob("data*.csv")
#creates lists of appropriate files (genes, sequences, mutations)

trinucs = [
        'ATA', 'ATC', 'ATT', 'ATG',
        'ACA', 'ACC', 'ACG', 'ACT',
        'AAC', 'AAT', 'AAA', 'AAG',
        'AGC', 'AGT', 'AGA', 'AGG',                 
        'CTA', 'CTC', 'CTG', 'CTT',
        'CCA', 'CCC', 'CCG', 'CCT',
        'CAC', 'CAT', 'CAA', 'CAG',
        'CGA', 'CGC', 'CGG', 'CGT',
        'GTA', 'GTC', 'GTG', 'GTT',
        'GCA', 'GCC', 'GCG', 'GCT',
        'GAC', 'GAT', 'GAA', 'GAG',
        'GGA', 'GGC', 'GGG', 'GGT',
        'TCA', 'TCC', 'TCG', 'TCT',
        'TTC', 'TTT', 'TTA', 'TTG',
        'TAC', 'TAT', 'TAA', 'TAG',
        'TGC', 'TGT', 'TGA', 'TGG',
]
#creates list of all 64 trinucs

foster_dict = {tri: {"p1": None, "p2": None, "p3": None} for tri in trinucs}
wei_dict = {tri: {"p1": None, "p2": None, "p3": None} for tri in trinucs}
#initiates empty dictionary to store values 

def reverse_complement(sequence):
    comp = {"A": "T", "T": "A", "C": "G", "G": "C"}
    rev_seq = reversed(sequence)
    good_seq = []
    for b in rev_seq:
        good_b = comp[b]
        good_seq.append(good_b)
    good_seq = "".join(good_seq)
    return good_seq
#defines function for creating a reverse complement of the sequence, to be used for - strand genes 

pos_list = [1,2,3]

f_cnt = 0
#file counter

for f in genes_f:
    #unpack appropriate files using counter 

    dict_tri2frq1 = {}
    dict_tri2frq2 = {}
    dict_tri2frq3 = {}
    #creates dictionaries for trinuc freq for different positions 

    dict_tri2muts1 = {}
    dict_tri2muts2 = {}
    dict_tri2muts3 = {}
    #creates dictionaries for trinuc to mut freq for different positions 

    dict_tri2rate1 = {}
    dict_tri2rate2 = {}
    dict_tri2rate3 = {}
    #creates dictionaries for trinuc to mut rate for different positions 

    dict_tri2diff12 = {}
    dict_tri2diff13 = {}
    dict_tri2diff23 = {}
    #creates dictionaries for differences in rates

    all_dicts = [dict_tri2frq1, dict_tri2frq2, dict_tri2frq3,
                 dict_tri2muts1, dict_tri2muts2, dict_tri2muts3,
                 dict_tri2rate1,  dict_tri2rate2, dict_tri2rate3,
                 dict_tri2diff12, dict_tri2diff13, dict_tri2diff23]
    #adds all dictionary names into a list

    for d in all_dicts:
        for t in trinucs:
            d[t] = 0
    #adds a key for each trinuc into all dictionaries and associates them with 0

    #now all dictionaries are set up to store information

    f_o = open(f, "r")
    all_genes = f_o.read()
    f_o.close()
    #opens the genes file

    seq_o = open(seq_f[f_cnt], "r")
    all_seq = seq_o.read()
    seq_o.close()
    #opens the appropriate sequence file 

    mut_o = open(mut_f[f_cnt], "r")
    all_mut = mut_o.read()
    mut_o.close()
    #opens the appropriate mutations file

    #now all our relevant files are open for the iteration

    gene_lines = all_genes.splitlines()
    gene_lines = gene_lines[1:]
    mut_lines = all_mut.splitlines()
    mut_lines = mut_lines[1:]
    #splits lines and removes header from genes and mutation data

    seq_raw = all_seq.splitlines()
    seq_raw = "".join(seq_raw[1:])
    #splits lines of sequence then removes header and joins into one long string 
    whole_seq = re.sub('[^A-Za-z]',"", seq_raw)
    whole_seq = whole_seq.upper()
    #this ensures all items are letters, and that they are all upper case, just in case

    if f_cnt == 0:
        expect_genome_len = int(4639675)
        #length of .2 on webpage
    else: 
        expect_genome_len = int(4641652)
        #length of .3 from webpage
    #takes the expected genome length from the relevant webpages
    
    actual_genome_len = int(len(whole_seq))
    diff_len = expect_genome_len - actual_genome_len
    #calculates difference in length of genome (from webpage) and genome list
    print(f"the genome list in interation {f_cnt} is {diff_len} bp different than expected")
    #sanity check to ensure same length as actual genome


    mut_list = []

    for mut in mut_lines:
        bits = mut.split(",")
        position = int(bits[1])
        #extracts the positions from the data
        mut_list.append(position)
        #adds the position to the mutations list
    
    #now we have all the mutation positions in a list, this will allow easier searching through them
    cnt_cds = 0

    for gn in gene_lines:
        bits = gn.split(",")
        #splits line based on comma 
        genome = bits[0]
        loc_tag = bits[1]
        strt = int(bits[2])
        nd = int(bits[3])
        strand = bits[4]
        seq_type = bits[5]
        gene_names = bits[6]
        #saves all the annotations

        if loc_tag != "b2891":
            #skipping this loc tag as it is the only one not a multiple of 3
            if seq_type == "CDS":
                #filters out any non coding sequences such as tRNA
                cnt_cds += 1
                #increment cds counter 

                strt0 = strt - 1 
                nd0 = nd 
                #start and end corrected for python indexing that excludes the end coordinate 

                full_seq = whole_seq[strt0 -1 : nd0 + 1]
                #takes sequence from n to n

                if strand == "-":
                    full_seq = reverse_complement(full_seq)

                for pos in pos_list:
                    if pos == 1:
                        seq_strt_index = strt0 - 1
                        nd_index = nd0 - 1
                        #sets the start and end of the reading frame for analysis of position 1 mutations 
                    elif pos == 2:
                        seq_strt_index = strt0
                        nd_index = nd0
                        #sets the start and end of the reading frame for analysis of position 2 mutations 
                    elif pos == 3:
                        seq_strt_index = strt0 + 1
                        nd_index = nd0 + 1
                        #sets the start and end of the reading frame for analysis of position 3 mutations
                    #This logic shifts the reading frame based on which position is being analysed

                    seq = whole_seq[seq_strt_index:nd_index]
                    #creates a list of the gene sequence starting based on the focal position

                    if strand == "-":
                        seq = reverse_complement(seq)
                        #executes the reverse_complement function on - strands 

                    trinuc_seq = []
                    #creates a new list to store the sequence once it is split into trinucs

                    for i in range(0,len(seq),3):
                        trinuc_seq.append(seq[i:i+3])
                        #cycles through thw length of seq from index 0, in intervals of 3
                        #then appends into the trinuc_seq these trinucleotides 

                    for t in trinuc_seq:
                        if pos == 1:
                            dict_tri2frq1[t] += 1
                        elif pos == 2:
                            dict_tri2frq2[t] += 1
                        elif pos == 3:
                            dict_tri2frq3[t] += 1
                        #all trinucs already exist in the dictionaries so just need to increment the value associated with them
                    
                    #now we have counts of trinucleotides for each position being analysed 

                    mut_cnt = 0
                    #start mutation counter

                    for mut in mut_list:
                        if mut > strt0 and mut <= nd0: 
                            if strand == "-":
                                in_gene_pos = (mut - strt0) + 1
                            else:
                                in_gene_pos = (nd0 - mut) + 1
                            #calculates the position of the mutation within the focal gene
                            pos = in_gene_pos % 3
                            #takes remainder when dividing by 3
                            if pos == 0:
                                pos = 3
                            
                            focal_trinuc = full_seq[in_gene_pos - 1 : in_gene_pos + 2]
                            #picks out the trinuc we are in 

                            if pos == 1:
                                dict_tri2muts1[focal_trinuc] += 1 
                            elif pos == 2:
                                dict_tri2muts2[focal_trinuc] += 1
                            elif pos == 3:
                                dict_tri2muts3[focal_trinuc] += 1
                            #increments the mutation counter according to the position being analysed
                            
                            mut_list.pop(mut_cnt)
                            #removes mut item from the mut list

                        mut_cnt += 1
                    
                    #now we have counts of mutations per trinucleotide type for each position being analysed 

    print(f"There are {len(mut_list)} mutations not in CDS in iteration {f_cnt}")

    #here we will calculate the values for the rates dictionaries, can also write them to an outfile?

    dict_cnt = 1
    #dictionary counter 

    for d in all_dicts[3:6]:
        #selects the tri2muts dictionaries to iterate through
        if dict_cnt == 1:
            freqs = dict_tri2frq1
            final_dict = dict_tri2rate1
        elif dict_cnt == 2:
            freqs = dict_tri2frq2
            final_dict = dict_tri2rate2
        elif dict_cnt == 3:
            freqs = dict_tri2frq3
            final_dict = dict_tri2rate3
        #associated the correct frequencies dictionary

        for tri, muts in d.items():
            #iterating through the trinucleotides in each muts dictionary
            freq = freqs[tri]
            #gets the frequency of the trinucleotide from other dict
            muts_rate = muts / freq
            #calculates the rate of mutations for each position accoring to frequency of trinuc 
            final_dict[tri] += muts_rate
            #adds the mutation rate to the dictionary 
        
        dict_cnt += 1
        #increment the dictionary counter 
    
    #now we have the 64-element dictionaries that store the rate of mutations per trinucleotide per position
    
    #print(f"{dict_tri2rate1}")
    #print(f"{dict_tri2rate2}")
    #print(f"{dict_tri2rate3}")
    #just to check 
    
    print(cnt_cds)

    if f_cnt == 0:
        #write to foster vector files 
        for trinuc, rate in dict_tri2rate1.items():
            rates_foster_pos1.write(f"{trinuc},{dict_tri2muts1[trinuc]},{dict_tri2frq1[trinuc]},{rate}\n")
            foster_dict[trinuc]["p1"] = rate
        #writes first poition
        for trinuc, rate in dict_tri2rate2.items():
            rates_foster_pos2.write(f"{trinuc},{dict_tri2muts2[trinuc]},{dict_tri2frq2[trinuc]},{rate}\n")
            foster_dict[trinuc]["p2"] = rate
        #writes second position
        for trinuc, rate in dict_tri2rate3.items():
            rates_foster_pos3.write(f"{trinuc},{dict_tri2muts3[trinuc]},{dict_tri2frq3[trinuc]},{rate}\n")
            foster_dict[trinuc]["p3"] = rate
        #writes third position
    else: 
        #write to wei vector files
        #write to foster vector files 
        for trinuc, rate in dict_tri2rate1.items():
            rates_wei_pos1.write(f"{trinuc},{dict_tri2muts1[trinuc]},{dict_tri2frq1[trinuc]},{rate}\n")
            wei_dict[trinuc]["p1"] = rate
        #writes first poition
        for trinuc, rate in dict_tri2rate2.items():
            rates_wei_pos2.write(f"{trinuc},{dict_tri2muts2[trinuc]},{dict_tri2frq2[trinuc]},{rate}\n")
            wei_dict[trinuc]["p2"] = rate
        #writes second position
        for trinuc, rate in dict_tri2rate3.items():
            rates_wei_pos3.write(f"{trinuc},{dict_tri2muts3[trinuc]},{dict_tri2frq3[trinuc]},{rate}\n")
            wei_dict[trinuc]["p3"] = rate
        #writes third position

    pos1_list = np.array(list(dict_tri2rate1.values()))
    pos2_list = np.array(list(dict_tri2rate2.values()))
    pos3_list = np.array(list(dict_tri2rate3.values()))
    #put the rates into arryas for checking normality

    pos_rates = [pos1_list, pos2_list, pos3_list]
    #put arrays into list

    pos_cnt = 1
    #counter for positions 

    colours = ["skyblue", "purple", "pink"]

    """
    for array in pos_rates:
        stat, p = stats.shapiro(array)
        #does shapiro wilk test

        print(f"for position {pos_cnt}:")
        print("Shapiro-Wilk W:", stat)
        print("p-value:", p)
        if p > 0.05:
            print("Data are consistent with normality")
        else:
            print("Data significantly deviate from normality")
        
        colour = colours[pos_cnt % len(colours)]
        #selects the right colout

        fig, ax = plt.subplots(1, 2, figsize=(12, 5))
        #creates figures space, 1 row, 2 columns, adjusts figure size to fit 2 plots

        ax[0].hist(array, bins=10, color=colour, edgecolor="black")
        ax[0].set_xlabel("Mutation Rate")
        ax[0].set_ylabel("Count")

        stats.probplot(array, dist="norm", plot=ax[1])

        if f_cnt == 0:
            ax[0].set_title(f"Foster, Position {pos_cnt}, Histohgram")
            ax[1].set_title(f"Foster, Position {pos_cnt}, Q-Q Plot")
        else: 
            ax[0].set_title(f"Wei, Position {pos_cnt}, Histohgram")
            ax[1].set_title(f"Wei, Position {pos_cnt}, Q-Q Plot")
        #histogram and qq plot to visualise the normality of the data
        
        plt.tight_layout()
        plt.show()

        pos_cnt += 1
    """
        
    #compare AAA rate in 2 to 3 and 1 to 3

    #need to make a vector that takes away one from the other 

    for trinuc, rate in dict_tri2rate3.items():
        rate_1 = dict_tri2rate1[trinuc]
        rate_2 = dict_tri2rate2[trinuc]
        #retrieves the rates of the same trinuc for the two other positions 

        diff_12 = rate_1 - rate_2
        diff_13 = rate - rate_1
        diff_23 = rate - rate_2
        #calculates the differneces in rates 

        dict_tri2diff12[trinuc] = diff_12
        dict_tri2diff13[trinuc] = diff_13
        dict_tri2diff23[trinuc] = diff_23
        #changes the values in the differences dictionaries (currently 0) into the calculated difference
    
    diff_12_array = np.array(list(dict_tri2diff12.values()))
    diff_13_array = np.array(list(dict_tri2diff13.values()))
    diff_23_array = np.array(list(dict_tri2diff23.values()))
    #inputs the differeneces into arrays

    diff_arrays = [diff_12_array, diff_13_array, diff_23_array]
    #makes list of array names

    diff_cnt1 = 1
    diff_cnt2 = 2
    #starts differences counetr

    for array in diff_arrays:
        stat, p = stats.shapiro(array)
        #does shapiro wilk test

        print(f"for difference between {diff_cnt1} and {diff_cnt2}:")
        print("testing for normality:")
        print("Shapiro-Wilk W:", stat)
        print("p-value:", p)
        if p > 0.05:
            print("Data are consistent with normality")
        else:
            print("Data significantly deviate from normality")
        
        #only one dataset deviates from normality (wei between 2 and 3)
        #therefore non-parametric tests need to be undertaken
        #this will be a wilcoxon test, the non-parametric version of the paired t-test

        stat2, p2 = wilcoxon(array)
        #defaults to two-sided test
        print("hypothesis testing:")
        print(f"Wilcoxon V-Statistic: {stat2}")
        print(f"p-value: {p2}")
        #prints out test values 

        if p2 > 0.05:
            print("Data is not significantly different from 0")
        else:
            print("Data is significantly different from 0")


        if f_cnt == 0:
            out_f.write(f"foster,{diff_cnt1} and {diff_cnt2},{stat2},{p2}\n")
            #writes stats to foster file
        else:
            out_f.write(f"wei,{diff_cnt1} and {diff_cnt2},{stat2},{p2}\n")
            #writes stats to wei file
    

        if diff_cnt2 != 2: 
            diff_cnt1 += 1
        
        if diff_cnt1 == 1:
            diff_cnt2 += 1
        #increment differences counters

    f_cnt += 1
    #increment file counter


for trinuc in trinucs:
    foster_1 = float(foster_dict[trinuc]["p1"])
    foster_2 = float(foster_dict[trinuc]["p2"])
    foster_3 = float(foster_dict[trinuc]["p3"])

    wei_1 = float(wei_dict[trinuc]["p1"])
    wei_2 = float(wei_dict[trinuc]["p2"])
    wei_3 = float(wei_dict[trinuc]["p3"])

    #finds all rates for specific trinuc in dicts#

    foster_p1 = float(f"{foster_1 * 1000:.3f}")
    foster_p2 = float(f"{foster_2 * 1000:.3f}")
    foster_p3 = float(f"{foster_3 * 1000:.3f}")

    wei_p1 = float(f"{wei_1 * 1000:.3f}")
    wei_p2 = float(f"{wei_2 * 1000:.3f}")
    wei_p3 = float(f"{wei_3 * 1000:.3f}")
    
    #times all the values by 1000 so they look good in table (headers indicate this)
    #also rounds to three decimals, no number is smaller than 3 decimal places so this is safe. 

    ws.append([trinuc, foster_p1, foster_p2, foster_p3, wei_p1, wei_p2, wei_p3])
    #writes to the excel spreadsheet

for col in ws.columns:
    max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    ws.column_dimensions[col[0].column_letter].width = max_len + 2
    #automatically sets column dimensions

wb.save("formatted_rates_table.xlsx")
#saves the excel file

for f in rates_f:
    f.close()
#closes all vector files

out_f.close()
#closes stats outfile

op_time = time.time() - strt_time
print(f"the operation took {op_time} seconds")


